"""
通用数据处理模块

提供数据清洗、类型转换、格式化等通用数据处理功能。
"""

import json
import logging
import os
import re
import shutil
import tempfile
import warnings
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from chatdb.utils.logger import logger


class DataProcessor:
    """通用数据处理器"""

    def remove_empty_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """删除完全为空的列"""
        df_cleaned = df.dropna(axis=1, how="all")

        empty_cols = [
            col
            for col in df_cleaned.columns
            if df_cleaned[col]
            .apply(lambda x: pd.isna(x) or (isinstance(x, str) and x.strip() == ""))
            .all()
        ]

        if empty_cols:
            df_cleaned = df_cleaned.drop(columns=empty_cols)

        return df_cleaned

    def remove_duplicate_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """删除列名和数据值都完全重复的列"""
        columns_to_remove = []

        for i, col1 in enumerate(df.columns):
            if col1 in columns_to_remove:
                continue

            for col2 in df.columns[i + 1 :]:
                if col2 in columns_to_remove or col1 != col2:
                    continue

                try:
                    if df[col1].equals(df[col2]):
                        columns_to_remove.append(col2)
                except Exception:
                    try:
                        if (
                            df[col1]
                            .fillna("__NULL__")
                            .equals(df[col2].fillna("__NULL__"))
                        ):
                            columns_to_remove.append(col2)
                    except Exception:
                        pass

        if columns_to_remove:
            df_cleaned = df.drop(columns=columns_to_remove)
        else:
            df_cleaned = df

        return df_cleaned

    def format_date_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """格式化日期列为YYYY-MM-DD格式"""
        df_formatted = df.copy()

        for col in df_formatted.columns:
            if pd.api.types.is_datetime64_any_dtype(df_formatted[col]):
                df_formatted[col] = df_formatted[col].apply(
                    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None
                )
            elif df_formatted[col].dtype == "object":
                non_null_values = df_formatted[col].dropna()
                if len(non_null_values) > 0:
                    sample_val = non_null_values.iloc[0]
                    if isinstance(sample_val, (pd.Timestamp, datetime)):
                        df_formatted[col] = df_formatted[col].apply(
                            lambda x: x.strftime("%Y-%m-%d")
                            if pd.notna(x)
                            and isinstance(x, (pd.Timestamp, datetime))
                            else x
                        )

        return df_formatted

    def convert_id_columns_to_string(
        self, df: pd.DataFrame, id_columns: List[str]
    ) -> pd.DataFrame:
        """将 ID 列转换为字符串类型"""
        df_converted = df.copy()

        def convert_to_str(x):
            if pd.isna(x):
                return None
            if isinstance(x, float) and x == int(x):
                return str(int(x))
            return str(x)

        for col in id_columns:
            if col in df_converted.columns:
                try:
                    df_converted[col] = (
                        df_converted[col].apply(convert_to_str).astype("object")
                    )
                except Exception as e:
                    logger.warning(f"转换ID列 '{col}' 失败: {e}")

        return df_converted

    def convert_column_types(
        self, df: pd.DataFrame, id_columns: List[str] = None
    ) -> pd.DataFrame:
        """
        根据字段的实际值进行智能类型转换

        转换策略：
        1. 尝试转换为日期类型
        2. 尝试转换为数值类型（整数或浮点数）
        3. 如果都失败，保持字符串类型

        Args:
            df: 原始DataFrame
            id_columns: ID 列名列表，这些列跳过数值转换

        Returns:
            转换后的DataFrame
        """
        if id_columns is None:
            id_columns = []

        df_converted = df.copy()

        for col in df_converted.columns:
            # 跳过 ID 列
            if col in id_columns:
                continue

            # 跳过已经是数值类型的列
            if df_converted[col].dtype in ["int64", "float64", "int32", "float32"]:
                continue

            # 跳过已经是日期类型的列
            if pd.api.types.is_datetime64_any_dtype(df_converted[col]):
                continue

            # 只处理object类型的列
            if df_converted[col].dtype == "object":
                non_null_data = df_converted[col].dropna()

                if len(non_null_data) == 0:
                    continue

                # 策略: 尝试转换为数值类型（不再自动转换日期，避免数值被错误解析为1970-01-01）
                numeric_success_count = 0
                has_decimal = False
                
                sample_size = min(100, len(non_null_data))
                sample_data = non_null_data.head(sample_size)

                try:
                    for val in sample_data:
                        if pd.notna(val):
                            val_str = str(val).strip()
                            val_str = (
                                val_str.replace(",", "")
                                .replace("￥", "")
                                .replace("$", "")
                                .replace("€", "")
                                .replace(" ", "")
                            )

                            try:
                                float_val = float(val_str)
                                numeric_success_count += 1
                                if "." in str(val) and float_val != int(float_val):
                                    has_decimal = True
                            except Exception:
                                pass

                    if numeric_success_count > sample_size * 0.5:
                        try:
                            df_converted[col] = df_converted[col].astype(str)
                            df_converted[col] = (
                                df_converted[col]
                                .str.replace(",", "")
                                .str.replace("￥", "")
                                .str.replace("$", "")
                                .str.replace("€", "")
                                .str.strip()
                            )
                            df_converted[col] = pd.to_numeric(
                                df_converted[col], errors="coerce"
                            )

                            if not has_decimal and df_converted[col].notna().any():
                                all_integers = True
                                for val in df_converted[col].dropna():
                                    if pd.notna(val) and val != int(val):
                                        all_integers = False
                                        break

                                if all_integers:
                                    df_converted[col] = df_converted[col].astype("Int64")
                                    logger.debug(f"列 '{col}' 转换为整数类型")
                                else:
                                    df_converted[col] = df_converted[col].astype(
                                        "float64"
                                    )
                                    logger.debug(f"列 '{col}' 转换为浮点数类型")
                            else:
                                df_converted[col] = df_converted[col].astype("float64")
                                logger.debug(f"列 '{col}' 转换为浮点数类型")

                            continue
                        except Exception as e:
                            logger.debug(f"列 '{col}' 转换为数值类型失败: {e}")
                            pass
                except Exception as e:
                    logger.debug(f"列 '{col}' 数值类型检测失败: {e}")
                    pass

        return df_converted

    def format_numeric_columns(
        self, df: pd.DataFrame, id_columns: List[str] = None
    ) -> pd.DataFrame:
        """
        格式化数值列为两位小数

        Args:
            df: 原始DataFrame
            id_columns: ID 列名列表，这些列跳过格式化

        Returns:
            格式化后的DataFrame（数值列保留两位小数）
        """
        if id_columns is None:
            id_columns = []

        df_formatted = df.copy()

        for col in df_formatted.columns:
            # 跳过 ID 列
            if col in id_columns:
                continue

            # 只处理数值类型的列
            if df_formatted[col].dtype in [
                "int64",
                "float64",
                "int32",
                "float32",
                "Int64",
            ]:
                try:
                    df_formatted[col] = pd.to_numeric(
                        df_formatted[col], errors="coerce"
                    ).apply(lambda x: round(float(x), 2) if pd.notna(x) else x)
                    df_formatted[col] = df_formatted[col].astype("float64")
                    logger.debug(f"列 '{col}' 格式化为两位小数 (float64)")
                except Exception as e:
                    logger.warning(f"格式化列 '{col}' 失败: {e}")
                    pass

        return df_formatted

    def get_sample_rows(self, df: pd.DataFrame, n: int = 2) -> List[Dict[str, Any]]:
        """
        获取n行样本数据，用于query改写时提供真实数据参考

        Args:
            df: DataFrame对象
            n: 样本行数

        Returns:
            样本数据列表，每行是一个字典
        """
        try:
            sample_df = df.head(n)
            rows = []
            for _, row in sample_df.iterrows():
                row_dict = {}
                for col in sample_df.columns:
                    val = row[col]
                    if pd.isna(val):
                        row_dict[col] = None
                    elif isinstance(val, (pd.Timestamp, datetime)):
                        row_dict[col] = (
                            val.strftime("%Y-%m-%d %H:%M:%S")
                            if hasattr(val, "strftime")
                            else str(val)
                        )
                    elif isinstance(val, (int, float)):
                        row_dict[col] = val
                    else:
                        row_dict[col] = str(val)
                rows.append(row_dict)
            return rows
        except Exception as e:
            logger.warning(f"获取样本数据失败: {e}")
            return []

    def convert_to_json_serializable(self, obj: Any) -> Any:
        """
        递归转换对象为可JSON序列化的格式

        Args:
            obj: 要转换的对象

        Returns:
            可JSON序列化的对象
        """
        from datetime import date, datetime

        if isinstance(obj, (datetime, date, pd.Timestamp)):
            if isinstance(obj, pd.Timestamp):
                if obj.hour == 0 and obj.minute == 0 and obj.second == 0:
                    return obj.strftime("%Y-%m-%d")
                else:
                    return obj.isoformat()
            elif isinstance(obj, datetime):
                if obj.hour == 0 and obj.minute == 0 and obj.second == 0:
                    return obj.strftime("%Y-%m-%d")
                else:
                    return obj.isoformat()
            elif isinstance(obj, date):
                return obj.strftime("%Y-%m-%d")
            else:
                return obj.isoformat()
        elif isinstance(obj, dict):
            return {k: self.convert_to_json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [self.convert_to_json_serializable(item) for item in obj]
        elif pd.isna(obj):
            return None
        elif isinstance(obj, (int, float, str, bool, type(None))):
            return obj
        else:
            return str(obj)

    def get_table_preview_data(
        self, db_path: str, table_name: str, limit: int = None, file_name: str = None
    ) -> Dict[str, Any]:
        """
        从数据库中获取表格预览数据

        Args:
            db_path: 数据库路径
            table_name: 表名
            limit: 获取的行数限制，None表示不限制
            file_name: 文件名（可选），用于在前端显示

        Returns:
            包含列定义和数据行的字典
        """
        import duckdb

        try:
            conn = duckdb.connect(db_path, read_only=True)

            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            columns = [
                {"field": col[0], "type": col[1], "headerName": col[0]}
                for col in columns_result
            ]

            total_count = conn.execute(
                f'SELECT COUNT(*) FROM "{table_name}"'
            ).fetchone()[0]

            if limit is None:
                rows_result = conn.execute(f'SELECT * FROM "{table_name}"').fetchall()
            else:
                rows_result = conn.execute(
                    f'SELECT * FROM "{table_name}" LIMIT {limit}'
                ).fetchall()

            rows = []
            for idx, row in enumerate(rows_result):
                row_dict = {"id": idx}
                for col_idx, col_info in enumerate(columns):
                    value = row[col_idx]
                    row_dict[col_info["field"]] = self.convert_to_json_serializable(
                        value
                    )
                rows.append(row_dict)

            conn.close()

            result = {"columns": columns, "rows": rows, "total": total_count}

            if file_name:
                result["file_name"] = file_name

            return result

        except Exception as e:
            logger.error(f"获取表格预览数据失败: {e}")
            return {"columns": [], "rows": [], "total": 0, "error": str(e)}

    def generate_create_table_sql(self, db_path: str, table_name: str) -> str:
        """
        生成建表SQL语句

        Args:
            db_path: 数据库路径
            table_name: 表名

        Returns:
            建表SQL语句
        """
        import duckdb

        try:
            conn = duckdb.connect(db_path, read_only=True)
            columns_result = conn.execute(f'DESCRIBE "{table_name}"').fetchall()
            conn.close()

            columns_sql = []
            for col in columns_result:
                col_name = col[0]
                col_type = col[1]
                columns_sql.append(f'    "{col_name}" {col_type}')

            create_sql = (
                f'CREATE TABLE "{table_name}" (\n' + ",\n".join(columns_sql) + "\n);"
            )
            return create_sql
        except Exception as e:
            logger.error(f"生成建表SQL失败: {e}")
            return ""

    def process_excel_sheet(
        self,
        df_raw: pd.DataFrame,
        id_columns: List[str] = None,
    ) -> pd.DataFrame:
        """
        第一阶段：处理 Excel Sheet 数据，转换为 DuckDB 可导入的形式
        
        包括：
        1. 数据清洗（删除空列、重复列、格式化日期）
        2. 数据类型转换（ID列转字符串、智能类型转换、格式化数值）
        3. 列名清理和去重
        
        Args:
            df_raw: 原始 DataFrame（已处理多级表头）
            id_columns: ID 列列表（可选）
            
        Returns:
            清洗和转换后的 DataFrame
        """
        if id_columns is None:
            id_columns = []
        
        # 1. 数据清洗
        df = self.remove_empty_columns(df_raw)
        df = self.remove_duplicate_columns(df)
        df = self.format_date_columns(df)
        
        # 2. 数据类型转换
        if id_columns:
            df = self.convert_id_columns_to_string(df, id_columns)
        df = self.convert_column_types(df, id_columns)
        df = self.format_numeric_columns(df, id_columns)
        
        # 3. 清理列名
        df.columns = [
            str(col)
            .replace(" ", "")
            .replace("\u00a0", "")
            .replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            for col in df.columns
        ]
        
        # 4. 去重列名
        final_columns = []
        seen_columns = {}
        for col in df.columns:
            if col in seen_columns:
                seen_columns[col] += 1
                final_columns.append(f"{col}_{seen_columns[col]}")
            else:
                seen_columns[col] = 0
                final_columns.append(col)
        df.columns = final_columns
        
        return df

    def import_to_duckdb(
        self,
        df: pd.DataFrame,
        db_path: str,
        table_name: str,
        id_columns: List[str] = None,
    ) -> None:
        """
        将清洗后的 DataFrame 导入到 DuckDB
        
        Args:
            df: 清洗后的 DataFrame
            db_path: DuckDB 数据库路径
            table_name: 表名
            id_columns: ID 列列表（可选）
        """
        import duckdb
        
        # 数据清理：删除空列和空字符串
        df_clean = df.copy()
        df_clean.replace("", pd.NA, inplace=True)
        
        # 删除所有列为空的 Unnamed 列
        unnamed_columns = [
            col for col in df_clean.columns
            if col.startswith("Unnamed") and df_clean[col].isnull().all()
        ]
        if unnamed_columns:
            df_clean.drop(columns=unnamed_columns, inplace=True)
        
        # 类型转换：确保所有列都是 DuckDB 可识别的 Python 原生类型
        # 注意：数值优先，不再尝试日期转换（日期转换已在 convert_column_types 中完成）
        for col in df_clean.columns:
            col_data = df_clean[col]
            
            # 全空列直接转为 object
            if col_data.isnull().all():
                df_clean[col] = col_data.astype('object')
                continue
            
            # 尝试转换为数值（数值优先，避免数值被错误解析为日期）
            try:
                numeric_series = pd.to_numeric(col_data, errors='coerce')
                if numeric_series.notna().sum() > len(col_data) * 0.3:
                    df_clean[col] = numeric_series.astype('float64')
                    continue
            except (ValueError, TypeError):
                pass
            
            # 字符串列转为 object 类型（避免 pandas StringDtype）
            df_clean[col] = col_data.astype('object')
        
        # 统一替换所有空值和无效字符串为 None
        df_clean = df_clean.replace({pd.NA: None, 'nan': None, 'None': None, '': None})
        
        # 转换为纯 Python 对象，避免 pandas 扩展类型问题
        records = df_clean.to_dict('records')
        df_python = pd.DataFrame(records)
        
        # 导入到 DuckDB
        conn = duckdb.connect(db_path)
        try:
            table_name_quoted = f'"{table_name}"'
            conn.execute(f"DROP TABLE IF EXISTS {table_name_quoted}")
            conn.execute(f"CREATE TABLE {table_name_quoted} AS SELECT * FROM df_python")
            logger.info(f"表 '{table_name}' 导入完成: {len(df_python)}行, {len(df_python.columns)}列")
        except Exception as e:
            logger.error(f"导入表 '{table_name}' 失败: {e}")
            raise
        finally:
            conn.close()

    # ==================== Excel 前处理方法（表头识别和筛选转换）====================

    def remove_excel_filters(self, excel_file_path: str) -> str:
        """去除 Excel 文件的筛选状态"""
        file_ext = Path(excel_file_path).suffix.lower()
        if file_ext == '.xls':
            return excel_file_path
        
        try:
            temp_dir = tempfile.mkdtemp()
            temp_xlsx = os.path.join(temp_dir, "temp_output.xlsx")
            filters_removed = False
            
            with zipfile.ZipFile(excel_file_path, 'r') as zip_in:
                with zipfile.ZipFile(temp_xlsx, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                    for item in zip_in.namelist():
                        data = zip_in.read(item)
                        
                        if item.startswith('xl/worksheets/sheet') and item.endswith('.xml'):
                            content = data.decode('utf-8')
                            original_content = content
                            
                            # 删除 autoFilter 和 hidden 属性
                            content = re.sub(r'<autoFilter[^>]*/>', '', content)
                            content = re.sub(r'<autoFilter[^>]*>.*?</autoFilter>', '', content, flags=re.DOTALL)
                            content = re.sub(r'(<row[^>]*)\s+hidden="1"([^>]*>)', r'\1\2', content)
                            content = re.sub(r'(<row[^>]*)\s+hidden="true"([^>]*>)', r'\1\2', content, flags=re.IGNORECASE)
                            
                            if content != original_content:
                                filters_removed = True
                            data = content.encode('utf-8')
                        
                        zip_out.writestr(item, data)
            
            if filters_removed:
                shutil.move(temp_xlsx, excel_file_path)
            else:
                os.remove(temp_xlsx)
            shutil.rmtree(temp_dir, ignore_errors=True)
            return excel_file_path
            
        except Exception as e:
            logger.warning(f"去除筛选失败，尝试openpyxl: {e}")
            return self._remove_excel_filters_openpyxl(excel_file_path)
    
    def _remove_excel_filters_openpyxl(self, excel_file_path: str) -> str:
        """使用 openpyxl 去除筛选状态（备用方案）"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_file_path)
            filters_removed = False
            
            for ws in wb.worksheets:
                if ws.auto_filter and ws.auto_filter.ref:
                    ws.auto_filter.ref = None
                    filters_removed = True
                for row in ws.row_dimensions:
                    if ws.row_dimensions[row].hidden:
                        ws.row_dimensions[row].hidden = False
                        filters_removed = True
            
            if filters_removed:
                wb.save(excel_file_path)
            wb.close()
            return excel_file_path
        except Exception as e:
            logger.warning(f"openpyxl去除筛选失败: {e}")
            return excel_file_path

    async def process_excel_file_with_header_detection(
        self,
        excel_file_path: str,
        sheet_name: str = None,
        llm_client=None,
        model_name: str = None,
    ) -> pd.DataFrame:
        """
        处理 Excel 文件，包括去除筛选、读取文件、检测表头、处理多级表头
        
        Args:
            excel_file_path: Excel 文件路径
            sheet_name: sheet 名称
            llm_client: LLM 客户端（可选，用于表头识别）
            model_name: 模型名称（可选）
            
        Returns:
            处理后的 DataFrame
        """
        from chatdb.database.base import BaseDatabaseConnector
        
        # 1. 去除筛选
        excel_file_path = self.remove_excel_filters(excel_file_path)
        
        # 2. 读取 Excel 文件（使用基类方法）
        df_raw = BaseDatabaseConnector.read_excel_file(excel_file_path, sheet_name=sheet_name, header=None)
        
        # 3. 处理多级表头
        df = await self.process_multi_level_header(df_raw, excel_file_path, sheet_name, llm_client, model_name)
        
        return df

    async def process_multi_level_header(
        self,
        df_raw: pd.DataFrame,
        excel_file_path: str,
        sheet_name: str = None,
        llm_client=None,
        model_name: str = None,
    ) -> pd.DataFrame:
        """
        处理多级表头
        
        Args:
            df_raw: 原始 DataFrame
            excel_file_path: Excel 文件路径
            sheet_name: sheet 名称
            llm_client: LLM 客户端（可选）
            model_name: 模型名称（可选）
        """
        header_rows, color_info = await self.detect_header_rows_with_color(
            excel_file_path, sheet_name, llm_client, model_name
        )

        if not header_rows:
            header_rows = [0]

        combined_headers = self.merge_headers_by_color(color_info)

        cleaned_headers = []
        for header in combined_headers:
            cleaned = str(header)
            parts = cleaned.split("-")
            valid_parts = [
                p
                for p in parts
                if "=" not in p
                and "@@" not in p
                and "@" not in p
                and p.strip()
                and p.strip() not in ["-", "_", ""]
            ]

            if valid_parts:
                cleaned = "-".join(valid_parts)
            else:
                cleaned = f"Column_{len(cleaned_headers)}"

            cleaned = (
                cleaned.replace("\n", "")
                .replace("\r", "")
                .replace("\t", "")
                .replace(" ", "")
                .replace("\u00a0", "")
            )
            cleaned = cleaned.replace("--", "-").replace("__", "_").strip("-_")

            if not cleaned or cleaned in ["-", "_"]:
                cleaned = f"Column_{len(cleaned_headers)}"

            cleaned_headers.append(cleaned)

        final_headers = []
        seen = {}
        for header in cleaned_headers:
            if header in seen:
                seen[header] += 1
                final_headers.append(f"{header}_{seen[header]}")
            else:
                seen[header] = 0
                final_headers.append(header)

        data_start_row = max(header_rows) + 1
        data_df = df_raw.iloc[data_start_row:].copy()

        if len(final_headers) > len(data_df.columns):
            final_headers = final_headers[: len(data_df.columns)]
        elif len(final_headers) < len(data_df.columns):
            for i in range(len(final_headers), len(data_df.columns)):
                final_headers.append(f"Column_{i}")

        data_df.columns = final_headers
        data_df = data_df.dropna(how="all").reset_index(drop=True)

        return data_df

    async def detect_header_rows_with_color(
        self,
        excel_file_path: str,
        sheet_name: str = None,
        llm_client=None,
        model_name: str = None,
    ) -> Tuple[List[int], Dict]:
        """
        使用颜色信息和LLM检测表头行

        Args:
            excel_file_path: Excel文件路径
            sheet_name: sheet名称，如果为None则使用active sheet
            llm_client: LLM 客户端（可选）
            model_name: 模型名称（可选）
        """
        from chatdb.database.base import BaseDatabaseConnector
        
        file_ext = Path(excel_file_path).suffix.lower()
        
        rows_data = []
        rows_colors = []
        max_cols = 0
        
        # .xls 文件不支持 openpyxl，跳过颜色检测，其他流程保持一致
        if file_ext == '.xls':
            logger.info("检测到 .xls 格式，跳过颜色检测")
            # 读取前几行数据
            df_preview = BaseDatabaseConnector.read_excel_file(excel_file_path, sheet_name=sheet_name, header=None)
            
            max_check_rows = min(20, len(df_preview))
            max_cols = len(df_preview.columns) if len(df_preview) > 0 else 0
            
            for i in range(max_check_rows):
                row_values = [str(v) if pd.notna(v) else "" for v in df_preview.iloc[i].tolist()]
                row_colors = [None] * len(row_values)  # .xls 不支持颜色检测，全部设为 None
                rows_data.append(row_values)
                rows_colors.append(row_colors)
        else:
            # .xlsx 文件尝试使用 openpyxl 读取颜色信息
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_file_path)
                ws = wb[sheet_name] if sheet_name else wb.active

                max_check_rows = min(20, ws.max_row)
                max_cols = ws.max_column

                for row_idx in range(1, max_check_rows + 1):
                    row_values = []
                    row_colors = []
                    for col_idx in range(1, max_cols + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell_value = self._get_cell_value(cell)
                        row_values.append(cell_value if cell_value is not None else "")
                        row_colors.append(self._get_cell_bg_color(cell))
                    rows_data.append(row_values)
                    rows_colors.append(row_colors)
            except Exception as e:
                # openpyxl 读取失败（可能是无效 XML），回退到 pandas 读取
                logger.warning(f"openpyxl 读取 .xlsx 文件失败，回退到 pandas: {e}")
                df_preview = BaseDatabaseConnector.read_excel_file(excel_file_path, sheet_name=sheet_name, header=None)
                
                max_check_rows = min(20, len(df_preview))
                max_cols = len(df_preview.columns) if len(df_preview) > 0 else 0
                
                rows_data = []
                rows_colors = []
                for i in range(max_check_rows):
                    row_values = [str(v) if pd.notna(v) else "" for v in df_preview.iloc[i].tolist()]
                    row_colors = [None] * len(row_values)  # 无法获取颜色，全部设为 None
                    rows_data.append(row_values)
                    rows_colors.append(row_colors)

        if llm_client:
            try:
                header_rows = await self.detect_header_rows_with_llm_and_color(
                    rows_data, rows_colors, llm_client, model_name
                )
                if header_rows:
                    color_info = {
                        "rows_data": rows_data,
                        "rows_colors": rows_colors,
                        "header_rows": header_rows,
                        "max_cols": max_cols,
                    }
                    return header_rows, color_info
            except Exception as e:
                logger.warning(f"  LLM检测失败: {e}")

        # 如果没有LLM或LLM检测失败，默认使用第0行作为表头
        header_rows = [0]
        color_info = {
            "rows_data": rows_data,
            "rows_colors": rows_colors,
            "header_rows": header_rows,
            "max_cols": max_cols,
        }
        return header_rows, color_info

    async def detect_header_rows_with_llm_and_color(
        self,
        rows_data: List[List],
        rows_colors: List[List],
        llm_client,
        model_name: str = None,
    ) -> List[int]:
        """
        使用LLM和颜色信息检测表头行

        Args:
            rows_data: 前20行的数据
            rows_colors: 前20行的颜色信息
            llm_client: LLM 客户端（BaseLLM 实例）
            model_name: 模型名称（可选）

        Returns:
            表头行的索引列表（从0开始）
        """
        from chatdb.llm.base import BaseLLM, Message, _extract_json_from_text

        if not isinstance(llm_client, BaseLLM):
            raise ValueError("llm_client 必须是 BaseLLM 实例")

        # 构建字典格式的数据表示（包含颜色信息）
        rows_dict = []
        for idx, (row_data, row_colors) in enumerate(
            zip(rows_data[:20], rows_colors[:20])
        ):
            # 只显示前10列数据
            row_dict = {
                "row_index": idx,
                "columns": {}
            }
            for col_idx, val in enumerate(row_data[:10]):
                col_key = f"列{col_idx + 1}"
                row_dict["columns"][col_key] = str(val) if val else ""
            
            # 统计颜色分布
            color_counts = {}
            for color in row_colors:
                if color:
                    color_counts[color] = color_counts.get(color, 0) + 1
            row_dict["color_info"] = (
                ", ".join(
                    [f"{color[:8]}({count}列)" for color, count in color_counts.items()]
                )
                if color_counts
                else "无背景色"
            )
            rows_dict.append(row_dict)
        
        rows_dict_str = json.dumps(rows_dict, ensure_ascii=False, indent=2)

        # 构建prompt
        prompt = f"""你是一个Excel数据分析专家。请分析以下Excel文件的前20行数据，判断哪些行是表头行（列名行）。

**核心判断原则**：
表头行的本质特征是：它包含**字段名/列名**（描述数据的属性），而不是数据值本身。

**判断标准（按优先级排序）**：
1. **语义判断（最重要）**：
   - 如果一行包含的是**字段名/属性名**（如"本币"、"汇率"、"姓名"、"金额"等描述性文字），而后续行包含的是**具体数据值**（如"CNY"、"1"、"张三"、"100"等），则第一行很可能是表头
   - 表头行通常：词语简短、具有描述性、表示数据的维度或属性
   - 数据行通常：包含具体数值、代码、日期等实际数据

2. **结构判断**：
   - 如果第一行看起来像列名，后续行是数据，即使没有背景色，第一行也应该是表头
   - 表头行通常在数据行之前
   - 表头行的单元格内容通常是文本性质的描述性词语

3. **可能有多级表头（多行表头）**，包括：
   - 分类标签行（如"基本信息"、"订单信息"等分类标题，通常只有少数列有内容）
   - 具体列名行（包含所有列的具体名称）
   - 多级表头时，应该包含所有相关的表头行，从分类标签行到最具体的列名行

4. **背景色（辅助判断，不是必要条件）**：
   - 表头行可能有特殊的背景色，但**没有背景色的行也可能是表头**
   - 不要因为缺少背景色就排除明显的列名行

5. **排除规则**：
   - 表头行之前可能有汇总信息行、说明行（如"请勿删除"、公式等），这些不是表头
   - 忽略包含"@@"、"@"、"="等公式标记的行，这些是Excel内部标记行
   - **忽略公式标记行之前的行**（如果第10行是公式标记行，第9行很可能是英文标题行，应该忽略）

6. **中英文重复标题处理（极其重要）**：
   - 如果发现连续的两行分别是英文列名和中文列名，且含义相同（如"Name"对应"中英文名"、"StaffID"对应"员工ID"），则**只保留中文行，跳过英文行**
   - 判断方法：如果一行全是英文单词（如Name, Department, StaffID），下一行或隔一行是对应的中文（如中英文名, 部门, 员工ID），则英文行是重复的，不应该包含在 header_rows 中
   - 优先选择包含中文列名的行作为表头

**重要：请按以下步骤生成结果**：
1. 首先在"reason"中**简要**说明你的判断理由（1-2句话即可）
2. 然后在"header_rows"中包含应该保留的表头行索引

返回JSON格式：
{{
  "reason": "简要判断理由",
  "header_rows": [行索引列表，从0开始]
}}

示例（中英文重复标题）：
第8行是"基本信息(BasicInfo)"（分类标签），第9行是"Name, StaffID, Department"（英文列名），第10行是公式标记行，第11行是"中英文名, 员工ID, 部门"（中文列名），则返回：
{{
  "reason": "第8行是分类标签。第9行是英文列名，第11行是对应的中文列名，二者含义相同，只保留中文行。第10行是公式标记行，忽略。",
  "header_rows": [8, 11]
}}

现在请分析以下数据（每行包含行号、列数据和颜色分布信息）：

{rows_dict_str}

返回JSON格式的结果："""

        # 调用LLM
        from chatdb.utils.logger import log_llm_interaction
        
        response = await llm_client.generate(
            [Message(role="user", content=prompt)], 
            temperature=0, 
            max_tokens=1000
        )

        # 记录 LLM 交互（输入输出封装在一起）
        log_llm_interaction(logger, "表头识别", prompt, response.content, max_prompt_chars=300, max_response_chars=500)

        if not response.content:
            return []

        # 解析JSON结果
        try:
            json_str = _extract_json_from_text(response.content)
            if not json_str:
                raise Exception("无法提取JSON内容")
            
            # 清理控制字符
            try:
                result = json.loads(json_str)
            except json.JSONDecodeError as e:
                # 清理控制字符
                cleaned_chars = []
                in_string = False
                escape_next = False
                
                for i, char in enumerate(json_str):
                    if escape_next:
                        cleaned_chars.append(char)
                        escape_next = False
                    elif char == '\\' and in_string:
                        cleaned_chars.append(char)
                        escape_next = True
                    elif char == '"' and (i == 0 or json_str[i-1] != '\\'):
                        in_string = not in_string
                        cleaned_chars.append(char)
                    elif in_string and ord(char) < 32 and char not in '\n\r\t':
                        cleaned_chars.append(' ')
                    else:
                        cleaned_chars.append(char)
                
                json_str_cleaned = ''.join(cleaned_chars)
                
                try:
                    result = json.loads(json_str_cleaned)
                except json.JSONDecodeError:
                    logger.error(f"JSON解析失败（已尝试清理控制字符）: {e}")
                    logger.error(f"JSON字符串前500字符: {json_str[:500]}")
                    raise e

            header_rows = result.get("header_rows", [])
            reason = result.get("reason", "")


            # 验证结果
            if isinstance(header_rows, list) and all(
                isinstance(x, int) for x in header_rows
            ):
                valid_rows = [r for r in header_rows if 0 <= r < len(rows_data)]
                if valid_rows:
                    return sorted(valid_rows)
                else:
                    logger.warning(f"  表头行索引无效: {header_rows}")
                    return None
            else:
                logger.warning(f"  表头行格式无效: {header_rows}")
                return None
        except json.JSONDecodeError as e:
            logger.warning(f"  JSON解析失败: {e}")
            return None
        except Exception as e:
            logger.error(f"  解析LLM结果时发生错误: {e}")
            return None

    def merge_headers_by_color(self, color_info: Dict) -> List[str]:
        """
        基于颜色信息合并表头

        策略：
        1. 对于每个表头行，识别同一颜色的列
        2. 如果同一颜色的列中只有一个单元格有值，用该值填充其他列
        3. 将多行表头合并为单行，用"-"连接

        Args:
            color_info: 颜色信息字典

        Returns:
            合并后的表头列表
        """
        rows_data = color_info["rows_data"]
        rows_colors = color_info["rows_colors"]
        header_rows = color_info["header_rows"]
        max_cols = color_info["max_cols"]

        # 提取表头行的数据和颜色
        header_data = [rows_data[i] for i in header_rows]
        header_colors = [rows_colors[i] for i in header_rows]

        # 对每一行表头，按颜色和位置连续性分组
        filled_headers = []
        for row_idx, (row_data, row_colors) in enumerate(
            zip(header_data, header_colors)
        ):
            color_position_groups = []
            current_group = None
            current_color = None

            for col_idx, (value, color) in enumerate(zip(row_data, row_colors)):
                if color:
                    if (
                        color == current_color
                        and current_group
                        and col_idx == current_group[-1][0] + 1
                    ):
                        current_group.append((col_idx, value))
                    else:
                        if current_group:
                            color_position_groups.append((current_color, current_group))
                        current_group = [(col_idx, value)]
                        current_color = color
                else:
                    if current_group:
                        color_position_groups.append((current_color, current_group))
                        current_group = None
                        current_color = None

            if current_group:
                color_position_groups.append((current_color, current_group))

            filled_row = list(row_data)
            for color, cells in color_position_groups:
                non_empty_values = [
                    (idx, val) for idx, val in cells if val and str(val).strip()
                ]

                if len(non_empty_values) == 1:
                    fill_value = non_empty_values[0][1]
                    for col_idx, _ in cells:
                        filled_row[col_idx] = fill_value
                elif len(non_empty_values) > 1:
                    non_empty_indices = sorted([idx for idx, _ in non_empty_values])

                    for i, (val_idx, val) in enumerate(
                        sorted(non_empty_values, key=lambda x: x[0])
                    ):
                        start_idx = val_idx
                        if i < len(non_empty_indices) - 1:
                            end_idx = non_empty_indices[i + 1]
                        else:
                            end_idx = max([idx for idx, _ in cells]) + 1

                        cells_to_fill = [
                            (idx, v) for idx, v in cells if start_idx <= idx < end_idx
                        ]
                        for col_idx, _ in cells_to_fill:
                            filled_row[col_idx] = val

            filled_headers.append(filled_row)

        # 合并多级表头
        combined_headers = []
        for col_idx in range(max_cols):
            # 收集该列的所有非空值（从上层到底层）
            col_values = []
            for row_idx in range(len(filled_headers)):
                val = filled_headers[row_idx][col_idx]
                if val and str(val).strip():
                    val_str = str(val).strip()
                    val_str = (
                        val_str.replace("\n", "")
                        .replace("\r", "")
                        .replace("\t", "")
                        .replace(" ", "")
                        .replace("\u00a0", "")
                    )
                    if not col_values or val_str != col_values[-1]:
                        col_values.append(val_str)

            # 用"-"连接多级表头
            if col_values:
                combined = "-".join(col_values)
            else:
                combined = f"Column_{col_idx}"

            combined_headers.append(combined)

        return combined_headers

    def _get_cell_value(self, cell) -> Optional[str]:
        """获取单元格值，处理公式"""
        if cell.value is None:
            return None

        if cell.data_type == "f":
            try:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cleaned = self._clean_excel_formula(cell.value)
                    return cleaned if cleaned else None
            except Exception as e:
                logger.warning(f"获取公式计算结果失败: {e}")

        value_str = str(cell.value)
        return (
            value_str.replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            .replace(" ", "")
        )

    def _get_cell_bg_color(self, cell) -> Optional[str]:
        """获取单元格背景色"""
        fill = cell.fill
        if fill.fill_type == "solid" or fill.fill_type == "patternFill":
            fg_color = fill.fgColor
            if fg_color.type == "rgb":
                rgb = fg_color.rgb
                if rgb and len(rgb) == 8:
                    return rgb[2:]
                return rgb
            elif fg_color.type == "indexed":
                return f"indexed_{fg_color.indexed}"
            elif fg_color.type == "theme":
                tint = fg_color.tint if fg_color.tint else 0
                return f"theme_{fg_color.theme}_{tint:.2f}"
        return None

    def _clean_excel_formula(self, text: str) -> str:
        """清理Excel公式和特殊字符"""
        if not text:
            return text

        text_str = str(text)

        if text_str.startswith("="):
            quoted_texts = re.findall(r'["\']([^"\']+)["\']', text_str)
            if quoted_texts:
                text_str = "".join(quoted_texts)
            else:
                cleaned = re.sub(r"[=&()]", "", text_str)
                cleaned = re.sub(
                    r"CHAR\s*\(\s*\d+\s*\)", "", cleaned, flags=re.IGNORECASE
                )
                cleaned = re.sub(
                    r"CONCATENATE\s*\([^)]*\)", "", cleaned, flags=re.IGNORECASE
                )
                cleaned = re.sub(r"[^a-zA-Z0-9\u4e00-\u9fff_]", "", cleaned)
                if not cleaned:
                    return ""
                text_str = cleaned

        if "@@" in text_str or text_str.startswith("@"):
            text_str = re.sub(r"@@[^\u4e00-\u9fff-]*", "", text_str)
            text_str = text_str.replace("@", "")
            if not text_str.strip():
                return ""

        text_str = (
            text_str.replace("\n", "")
            .replace("\r", "")
            .replace("\t", "")
            .replace(" ", "")
        )
        return text_str

